import csv
import os
from openpyxl import load_workbook

# Define the directory containing the .xlsm files
source_directory = 'source_data'

# Create a list to store values
serial_nums = []
cal_dates = []
sweep_dates = []

# Create a list to store values from rows 8 through 11
rows_8_to_11_values = []

# Iterate over each file in the source directory
for filename in os.listdir(source_directory):
    if filename.endswith('.xlsm'):
        file_path = os.path.join(source_directory, filename)

        # Load the workbook with data_only=True to get cell values
        workbook = load_workbook(file_path, data_only=True)

        # Get the first sheet name
        first_sheet_name = workbook.sheetnames[0]

        # Select the first sheet from the workbook
        sheet = workbook[first_sheet_name]

        # Get the value in cell B2 and store it in the list
        serial_nums.append(sheet['B2'].value)
        cal_dates.append(sheet['B3'].value)
        sweep_dates.append(sheet['B4'].value)

        # Get values from rows 8 through 11 and store them in the list
        rows_8_to_11 = []
        for row_num in range(8, 12):  # Include rows 8 through 11
            row_values = []
            for cell in sheet[row_num]:
                row_values.append(cell.value)
            rows_8_to_11.append(row_values)
        rows_8_to_11_values.append(rows_8_to_11)

# Print the values of cell B2 and rows 8 through 11 for each workbook
# for i, (serial, rows_values) in enumerate(zip(serial_nums, rows_8_to_11_values)):
#     print(f"File: {os.listdir(source_directory)[i]}")
#     print(f"Value in cell B2: {serial}")
#     print("Rows 8 through 11:")
#     for row_num, row in enumerate(rows_values, start=8):
#         print(f"Row {row_num}: {row}")
#     print()

# Create and open a CSV file for writing
with open('output.csv', 'w', newline='') as csvfile:
    csv_writer = csv.writer(csvfile)

    # Write cell B2 values to the CSV file
    csv_writer.writerow(['File', 'Serial', 'Cal', 'Sweep', 'Value 1', 'Value 2', 'Value 3', 'Value 4', 'Value 5', 'Value 6', 'Value 7', 'Value 8', 'Value 9' , 'Value 10', 'Value 11'])
    for filename, serial, cal_, sweep_, rows_values in zip(os.listdir(source_directory), serial_nums, cal_dates, sweep_dates, rows_8_to_11_values):
       for row in rows_values:
         csv_writer.writerow([filename, serial, cal_, sweep_] + [cell for cell in row])     

print("Output saved to 'output.csv'")
